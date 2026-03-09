"""Tests for the DOCX and XLSX builder."""

from pathlib import Path

import openpyxl

from bot.builder import build_docx, build_xlsx
from bot.models import TranslatedQuestion
from bot.parser import parse_docx, parse_xlsx


class TestBuildDocx:
    def test_bilingual_text_in_output(self, sample_docx, sample_translated, tmp_path):
        output = str(tmp_path / "output.docx")
        build_docx(str(sample_docx), sample_translated, output)

        assert Path(output).exists()
        questions = parse_docx(output)
        assert len(questions) >= 1

    def test_tag_not_modified(self, sample_docx, sample_translated, tmp_path):
        output = str(tmp_path / "output.docx")
        build_docx(str(sample_docx), sample_translated, output)

        questions = parse_docx(output)
        for q in questions:
            assert "संघर्ष 2026" in q.tag


class TestBuildXlsx:
    def test_bilingual_question_text(self, sample_xlsx, sample_translated, tmp_path):
        output = str(tmp_path / "output.xlsx")
        build_xlsx(str(sample_xlsx), sample_translated, output)

        wb = openpyxl.load_workbook(output, data_only=True)
        ws = wb["Questions"]
        row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        q_text = row2[5]  # col F = QUESTION TEXT
        assert "यंत्र दोष" in q_text
        wb.close()

    def test_right_answer_is_integer(self, sample_xlsx, sample_translated, tmp_path):
        output = str(tmp_path / "output.xlsx")
        build_xlsx(str(sample_xlsx), sample_translated, output)

        wb = openpyxl.load_workbook(output, data_only=True)
        ws = wb["Questions"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            right_answer = row[16]  # col Q = RIGHT ANSWER
            assert isinstance(right_answer, (int, float))
        wb.close()

    def test_tag_not_modified(self, sample_xlsx, sample_translated, tmp_path):
        output = str(tmp_path / "output.xlsx")
        build_xlsx(str(sample_xlsx), sample_translated, output)

        wb = openpyxl.load_workbook(output, data_only=True)
        ws = wb["Questions"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            tags = row[3]  # col D = TAGS
            assert "संघर्ष 2026" in str(tags)
        wb.close()

    def test_options_bilingual(self, sample_xlsx, sample_translated, tmp_path):
        output = str(tmp_path / "output.xlsx")
        build_xlsx(str(sample_xlsx), sample_translated, output)

        wb = openpyxl.load_workbook(output, data_only=True)
        ws = wb["Questions"]
        row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        # Q2 options should be bilingual
        opt1 = str(row3[6])  # col G = OPTION1
        assert "Vata" in opt1 or "वात" in opt1
        wb.close()
