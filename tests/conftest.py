"""Shared fixtures for the bilingual bot tests."""

import os
import tempfile
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import PatternFill, Font, Alignment

from bot.models import Question, TranslatedQuestion

SAMPLES_DIR = Path(__file__).parent / "samples"


def _make_docx_xml(questions: list[dict]) -> str:
    """Generate a minimal document.xml for testing."""
    paras = []

    paras.append('<w:p><w:r><w:rPr><w:rFonts w:ascii="Courier New"/></w:rPr>'
                 '<w:t>#English_directions</w:t></w:r></w:p>')

    for q in questions:
        num = q["num"]
        serial = q.get("serial", f"{num}/3")
        paras.append(f'<w:p><w:r><w:t>#Question {num} </w:t></w:r></w:p>')
        paras.append(f'<w:p><w:r><w:t>[{serial}] @AIPGETMADEEASY</w:t></w:r></w:p>')
        paras.append(f'<w:p><w:r><w:t>{q["question"]}</w:t></w:r></w:p>')
        paras.append('<w:p><w:r><w:t>#Options ###A</w:t></w:r></w:p>')
        paras.append(f'<w:p><w:r><w:t>{q["options"]["A"]}</w:t></w:r></w:p>')
        paras.append('<w:p><w:r><w:t>###B</w:t></w:r></w:p>')
        paras.append(f'<w:p><w:r><w:t>{q["options"]["B"]}</w:t></w:r></w:p>')
        paras.append('<w:p><w:r><w:t>###C</w:t></w:r></w:p>')
        paras.append(f'<w:p><w:r><w:t>{q["options"]["C"]}</w:t></w:r></w:p>')
        paras.append('<w:p><w:r><w:t>###D</w:t></w:r></w:p>')
        paras.append(f'<w:p><w:r><w:t>{q["options"]["D"]}</w:t></w:r></w:p>')
        paras.append('<w:p><w:r><w:t>###E</w:t></w:r></w:p>')
        paras.append('<w:p><w:r><w:t>#Correct_option</w:t></w:r></w:p>')
        paras.append(f'<w:p><w:r><w:t>{q["correct"]}</w:t></w:r></w:p>')
        paras.append('<w:p><w:r><w:t>#Solution</w:t></w:r></w:p>')
        paras.append(f'<w:p><w:r><w:t>{q["solution"]}</w:t></w:r></w:p>')
        paras.append('<w:p><w:r><w:t>#tag</w:t></w:r></w:p>')
        paras.append(f'<w:p><w:r><w:t>{q["tag"]}</w:t></w:r></w:p>')

    body_content = "\n".join(paras)
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:wpc="http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas"'
        ' xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"'
        ' xmlns:o="urn:schemas-microsoft-com:office:office"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
        ' xmlns:m="http://schemas.openxmlformats.org/officeDocument/2006/math"'
        ' xmlns:v="urn:schemas-microsoft-com:vml"'
        ' xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"'
        ' xmlns:w10="urn:schemas-microsoft-com:office:word"'
        ' xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"'
        ' xmlns:wne="http://schemas.microsoft.com/office/word/2006/wordml">'
        f'<w:body>{body_content}</w:body></w:document>'
    )


SAMPLE_QUESTIONS = [
    {
        "num": 1,
        "serial": "1/3",
        "question": "How many Yantra dosha?",
        "options": {"A": "4", "B": "6", "C": "8", "D": "10"},
        "correct": "B",
        "solution": "Sushruta Sara pg 4, Su su 14/38",
        "tag": "Quiz no. 01 Date 1/8/2025 Topic - Sushruta brief notes 1-8 संघर्ष 2026 TEST SERIES",
    },
    {
        "num": 2,
        "serial": "2/3",
        "question": "त्वक् विवर्णता is the lakshana of",
        "options": {"A": "वात", "B": "पित्त", "C": "कफ", "D": "रक्त"},
        "correct": "A",
        "solution": "Reference: Su su 21/18",
        "tag": "Quiz no. 01 Date 1/8/2025 Topic - Sushruta brief notes 1-8 संघर्ष 2026 TEST SERIES",
    },
    {
        "num": 3,
        "serial": "3/3",
        "question": "Laghu anna is indicated in",
        "options": {"A": "Jwara", "B": "Atisara", "C": "Gulma", "D": "All"},
        "correct": "D",
        "solution": "Charaka Samhita Su 25/40",
        "tag": "Quiz no. 01 Date 1/8/2025 Topic - Sushruta brief notes 1-8 संघर्ष 2026 TEST SERIES",
    },
]


@pytest.fixture
def sample_docx(tmp_path: Path) -> Path:
    """Create a sample DOCX file for testing."""
    xml_content = _make_docx_xml(SAMPLE_QUESTIONS)
    docx_path = tmp_path / "sample.docx"

    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '</Types>'
    )

    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
        '</Relationships>'
    )

    with zipfile.ZipFile(docx_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("word/document.xml", xml_content)

    return docx_path


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """Create a sample XLSX file for testing."""
    xlsx_path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"

    headers = [
        "S No.", "SUBJECT", "TOPIC", "TAGS", "QUESTION TYPE",
        "QUESTION TEXT", "OPTION1", "OPTION2", "OPTION3", "OPTION4",
        "OPTION5", "OPTION6", "OPTION7", "OPTION8", "OPTION9", "OPTION10",
        "RIGHT ANSWER", "EXPLANATION", "CORRECT MARKS", "NEGATIVE MARKS", "DIFFICULTY",
    ]
    ws.append(headers)

    for hdr_cell in ws[1]:
        hdr_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        hdr_cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"

    for q in SAMPLE_QUESTIONS:
        correct_int = {"A": 1, "B": 2, "C": 3, "D": 4}[q["correct"]]
        row = [
            q["num"], None, None, q["tag"], "SINGLECORRECT",
            q["question"], q["options"]["A"], q["options"]["B"],
            q["options"]["C"], q["options"]["D"],
            None, None, None, None, None, None,
            correct_int, q["solution"], 4, 1, "Medium",
        ]
        ws.append(row)
        ws.row_dimensions[ws.max_row].height = 80

    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


@pytest.fixture
def sample_questions() -> list[Question]:
    """Return parsed Question objects matching the sample files."""
    return [
        Question(
            num=q["num"],
            serial=q["serial"],
            question=q["question"],
            options=q["options"],
            correct_answer=q["correct"],
            solution=q["solution"],
            tag=q["tag"],
            serial_line=f"[{q['serial']}] @AIPGETMADEEASY",
        )
        for q in SAMPLE_QUESTIONS
    ]


@pytest.fixture
def sample_translated(sample_questions: list[Question]) -> list[TranslatedQuestion]:
    """Return mock-translated questions."""
    return [
        TranslatedQuestion(
            original=sample_questions[0],
            question_bilingual="How many Yantra dosha?\nयंत्र दोष की संख्या कितनी है?",
            options_bilingual={"A": "4", "B": "6", "C": "8", "D": "10"},
        ),
        TranslatedQuestion(
            original=sample_questions[1],
            question_bilingual="त्वक् विवर्णता (Discolouration of skin) is the lakshana of",
            options_bilingual={
                "A": "वात (Vata)",
                "B": "पित्त (Pitta)",
                "C": "कफ (Kapha)",
                "D": "रक्त (Rakta)",
            },
        ),
        TranslatedQuestion(
            original=sample_questions[2],
            question_bilingual="Laghu anna is indicated in\nलघु अन्न किसमें निर्दिष्ट है?",
            options_bilingual={
                "A": "Jwara (ज्वर)",
                "B": "Atisara (अतिसार)",
                "C": "Gulma (गुल्म)",
                "D": "All (सभी)",
            },
        ),
    ]
