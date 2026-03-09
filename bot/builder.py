"""Build bilingual DOCX and XLSX output files."""

import os
import re
import shutil
import tempfile
import zipfile
from pathlib import Path

from lxml import etree
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from bot.models import TranslatedQuestion

_SERIAL_RE = re.compile(r"^\[\d+/\d+\]\s*@\w+")

WORD_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
NSMAP = {"w": WORD_NS}


def _find_paragraphs(body: etree._Element) -> list[etree._Element]:
    return body.findall("w:p", NSMAP)


def _para_text(para: etree._Element) -> str:
    parts = []
    for run in para.findall(".//w:r", NSMAP):
        for child in run:
            tag = etree.QName(child).localname
            if tag == "t":
                parts.append(child.text or "")
            elif tag == "br":
                parts.append("\n")
    return "".join(parts)


def _set_para_text_bilingual(para: etree._Element, new_text: str) -> None:
    """Replace all text in a paragraph's runs with new_text, using <w:br/> for newlines."""
    runs = para.findall(".//w:r", NSMAP)
    if not runs:
        return

    # Keep first run, clear all others
    first_run = runs[0]
    for run in runs[1:]:
        run.getparent().remove(run)

    # Remove existing <w:t> and <w:br/> from first run, keep <w:rPr>
    for child in list(first_run):
        tag = etree.QName(child).localname
        if tag in ("t", "br"):
            first_run.remove(child)

    # Add new text with line breaks
    lines = new_text.split("\n")
    for idx, line in enumerate(lines):
        t_el = etree.SubElement(first_run, f"{{{WORD_NS}}}t")
        t_el.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
        t_el.text = line
        if idx < len(lines) - 1:
            etree.SubElement(first_run, f"{{{WORD_NS}}}br")


def build_docx(
    source_path: str,
    questions: list[TranslatedQuestion],
    output_path: str,
) -> None:
    """Build a bilingual DOCX by modifying the source file's XML."""
    tmpdir = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(source_path, "r") as zf:
            zf.extractall(tmpdir)

        doc_xml_path = os.path.join(tmpdir, "word", "document.xml")
        tree = etree.parse(doc_xml_path)
        root = tree.getroot()
        body = root.find(".//w:body", NSMAP)
        paragraphs = _find_paragraphs(body)

        for tq in questions:
            orig = tq.original
            if not orig.para_indices:
                continue

            bilingual_q = tq.question_bilingual
            if not bilingual_q or not bilingual_q.strip():
                bilingual_q = orig.question

            for para_idx in orig.para_indices:
                if para_idx >= len(paragraphs):
                    continue
                para = paragraphs[para_idx]
                text = _para_text(para).strip()

                # Never touch serial line, markers, or section headers
                if (
                    not text
                    or text.startswith("#")
                    or text.startswith("###")
                    or _SERIAL_RE.match(text)
                ):
                    continue

                # Match question text paragraph
                if text == orig.question.strip():
                    _set_para_text_bilingual(para, bilingual_q)
                    continue

                # Match option values
                for opt_key, opt_val in orig.options.items():
                    if text == opt_val.strip() and opt_key in tq.options_bilingual:
                        _set_para_text_bilingual(para, tq.options_bilingual[opt_key])
                        break

        tree.write(doc_xml_path, xml_declaration=True, encoding="UTF-8", standalone=True)

        # Rezip
        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for dirpath, _, filenames in os.walk(tmpdir):
                for fn in filenames:
                    full = os.path.join(dirpath, fn)
                    arcname = os.path.relpath(full, tmpdir)
                    zout.write(full, arcname)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(wrap_text=True)
DATA_ALIGN = Alignment(wrap_text=True, vertical="top")

COLUMN_MAP = {
    "A": "S No.",
    "B": "SUBJECT",
    "C": "TOPIC",
    "D": "TAGS",
    "E": "QUESTION TYPE",
    "F": "QUESTION TEXT",
    "G": "OPTION1",
    "H": "OPTION2",
    "I": "OPTION3",
    "J": "OPTION4",
    "K": "OPTION5",
    "L": "OPTION6",
    "M": "OPTION7",
    "N": "OPTION8",
    "O": "OPTION9",
    "P": "OPTION10",
    "Q": "RIGHT ANSWER",
    "R": "EXPLANATION",
    "S": "CORRECT MARKS",
    "T": "NEGATIVE MARKS",
    "U": "DIFFICULTY",
}


def build_xlsx(
    source_path: str,
    questions: list[TranslatedQuestion],
    output_path: str,
) -> None:
    """Build a bilingual XLSX by updating the source file."""
    wb = openpyxl.load_workbook(source_path)
    ws = wb["Questions"]

    # Build header index
    headers = {}
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1):
        if cell.value:
            headers[str(cell.value).strip().upper()] = col_idx

    q_col = headers.get("QUESTION TEXT")
    opt_cols = [headers.get(f"OPTION{i}") for i in range(1, 5)]

    q_map = {tq.original.num: tq for tq in questions}

    for row in ws.iter_rows(min_row=2):
        sno_cell = row[0]
        if sno_cell.value is None:
            continue
        sno = int(sno_cell.value)
        tq = q_map.get(sno)
        if tq is None:
            continue

        bilingual_q = tq.question_bilingual
        if not bilingual_q or not bilingual_q.strip():
            bilingual_q = tq.original.question

        if q_col:
            row[q_col - 1].value = bilingual_q
            row[q_col - 1].alignment = DATA_ALIGN

        opt_keys = ["A", "B", "C", "D"]
        for key, col in zip(opt_keys, opt_cols):
            if col and key in tq.options_bilingual:
                row[col - 1].value = tq.options_bilingual[key]
                row[col - 1].alignment = DATA_ALIGN

    wb.save(output_path)
    wb.close()
