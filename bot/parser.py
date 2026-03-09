"""Parse DOCX and XLSX quiz files into Question objects."""

import re
import zipfile
import tempfile
from pathlib import Path

from lxml import etree
import openpyxl

from bot.models import Question


WORD_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
NSMAP = {"w": WORD_NS}


def _para_text(para_el: etree._Element) -> str:
    """Extract text from a <w:p> element, replacing <w:br/> with \\n."""
    parts: list[str] = []
    for run in para_el.findall(".//w:r", NSMAP):
        for child in run:
            tag = etree.QName(child).localname
            if tag == "t":
                parts.append(child.text or "")
            elif tag == "br":
                parts.append("\n")
    return "".join(parts)


def parse_docx(path: str) -> list[Question]:
    """Parse a DOCX file and return a list of Question objects."""
    with zipfile.ZipFile(path, "r") as zf:
        with zf.open("word/document.xml") as f:
            tree = etree.parse(f)

    body = tree.getroot().find(".//w:body", NSMAP)
    if body is None:
        return []

    paragraphs = body.findall("w:p", NSMAP)
    texts = [_para_text(p) for p in paragraphs]

    questions: list[Question] = []
    i = 0

    while i < len(texts):
        text = texts[i].strip()

        question_match = re.match(r"#Question\s+(\d+)", text)
        if not question_match:
            i += 1
            continue

        q_num = int(question_match.group(1))
        para_indices = [i]
        i += 1

        # Next paragraph: serial line like "[1/25] @AIPGETMADEEASY"
        serial = ""
        serial_line = ""
        if i < len(texts):
            serial_match = re.match(r"\[(\d+/\d+)\]\s*@\w+", texts[i].strip())
            if serial_match:
                serial = serial_match.group(0).split("]")[0].lstrip("[")
                serial_line = texts[i].strip()
            else:
                # Fallback: just extract the N/M part
                num_match = re.match(r"\[(\d+/\d+)\]", texts[i].strip())
                if num_match:
                    serial = num_match.group(1)
                    serial_line = texts[i].strip()
            para_indices.append(i)
            i += 1

        # Collect question text until we hit #Options
        q_text_parts: list[str] = []
        serial_line_re = re.compile(r"^\[\d+/\d+\]\s*@\w+")
        while i < len(texts):
            t = texts[i].strip()
            if t.startswith("#Options") or t.startswith("###"):
                break
            # Skip serial lines that leaked into question body
            if t and not serial_line_re.match(t):
                q_text_parts.append(texts[i])
            para_indices.append(i)
            i += 1

        question_text = "\n".join(q_text_parts).strip()

        # Parse options: look for ###A, ###B, ###C, ###D, ###E
        options: dict[str, str] = {}
        current_option = None

        # The first line may be "#Options ###A" combined
        if i < len(texts):
            first_opt = texts[i].strip()
            opt_match = re.search(r"###([A-E])", first_opt)
            if opt_match:
                current_option = opt_match.group(1)
            para_indices.append(i)
            i += 1

        while i < len(texts):
            t = texts[i].strip()
            if t.startswith("#Correct_option"):
                break

            opt_match = re.match(r"###([A-E])", t)
            if opt_match:
                current_option = opt_match.group(1)
                remainder = re.sub(r"###[A-E]\s*", "", t).strip()
                if remainder and current_option:
                    options[current_option] = remainder
                para_indices.append(i)
                i += 1
                continue

            if current_option and current_option not in options:
                options[current_option] = texts[i].strip()
            elif current_option and current_option in options:
                options[current_option] += "\n" + texts[i].strip()

            para_indices.append(i)
            i += 1

        # Skip ###E marker if present (no option E content needed)
        if "E" in options and not options["E"]:
            del options["E"]

        # Correct option
        correct = ""
        if i < len(texts) and texts[i].strip().startswith("#Correct_option"):
            para_indices.append(i)
            i += 1
            if i < len(texts):
                correct = texts[i].strip()
                para_indices.append(i)
                i += 1

        # Solution
        solution_parts: list[str] = []
        if i < len(texts) and texts[i].strip().startswith("#Solution"):
            para_indices.append(i)
            i += 1
            while i < len(texts):
                t = texts[i].strip()
                if t.startswith("#tag"):
                    break
                solution_parts.append(texts[i])
                para_indices.append(i)
                i += 1

        solution = "\n".join(solution_parts).strip()

        # Tag
        tag = ""
        if i < len(texts) and texts[i].strip().startswith("#tag"):
            para_indices.append(i)
            i += 1
            if i < len(texts):
                tag = texts[i].strip()
                para_indices.append(i)
                i += 1

        questions.append(
            Question(
                num=q_num,
                serial=serial,
                question=question_text,
                options=options,
                correct_answer=correct,
                solution=solution,
                tag=tag,
                serial_line=serial_line,
                para_indices=para_indices,
            )
        )

    return questions


INT_TO_LETTER = {1: "A", 2: "B", 3: "C", 4: "D"}


def parse_xlsx(path: str) -> list[Question]:
    """Parse an XLSX file and return a list of Question objects."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Questions"]

    headers = {}
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=0):
        if cell.value:
            headers[str(cell.value).strip().upper()] = col_idx

    required = {"S NO.", "QUESTION TEXT", "OPTION1", "OPTION2", "OPTION3", "OPTION4",
                "RIGHT ANSWER", "EXPLANATION", "TAGS"}
    if not required.issubset(set(headers.keys())):
        missing = required - set(headers.keys())
        raise ValueError(f"Missing columns in XLSX: {missing}")

    questions: list[Question] = []
    total_rows = ws.max_row or 1

    for row in ws.iter_rows(min_row=2, values_only=False):
        cells = list(row)
        sno = cells[headers["S NO."]].value
        if sno is None:
            continue

        q_text = str(cells[headers["QUESTION TEXT"]].value or "")
        opt1 = str(cells[headers["OPTION1"]].value or "")
        opt2 = str(cells[headers["OPTION2"]].value or "")
        opt3 = str(cells[headers["OPTION3"]].value or "")
        opt4 = str(cells[headers["OPTION4"]].value or "")

        right_answer_raw = cells[headers["RIGHT ANSWER"]].value
        right_answer_int = int(right_answer_raw) if right_answer_raw else 1
        correct_letter = INT_TO_LETTER.get(right_answer_int, "A")

        explanation = str(cells[headers["EXPLANATION"]].value or "")
        tags = str(cells[headers["TAGS"]].value or "")

        total = total_rows - 1
        serial = f"{sno}/{total}" if total > 0 else str(sno)

        questions.append(
            Question(
                num=int(sno),
                serial=serial,
                question=q_text,
                options={"A": opt1, "B": opt2, "C": opt3, "D": opt4},
                correct_answer=correct_letter,
                solution=explanation,
                tag=tags,
                right_answer_int=right_answer_int,
            )
        )

    wb.close()
    return questions
